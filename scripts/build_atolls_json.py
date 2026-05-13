"""Join lat/lon + L4 + L5 stats + L5 code lookup into one canonical atolls.json.

Applies the documented L4 unit fix (Taka Garlarang, Cato Reef are in m² not km²
in the source L4 sheet). Uses L5 'Total général' as canonical area when present.
"""
from __future__ import annotations

import csv
import difflib
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path("/storage/emulated/0/Download/coral-atoll")
STATS_DIR = ROOT / "references/dataverse_files/MCRMP/Statistics_Documentation"
LATLON = STATS_DIR / "Lat-Long_atolls-598.csv"
STATS = STATS_DIR / "Atolls_Statistics-598.xlsx"
CODES = STATS_DIR / "Millennium-Codes-Reefbase-2023-Atoll-Codes.xlsx"
OUT = ROOT / "docs/data/atolls.json"
OUT_PROCESSED = ROOT / "data/processed/atolls.json"

L4_UNIT_FIX = {"Taka Garlarang", "Cato Reef"}  # rows stored in m², divide by 1e6


def _f(s: str | None) -> float | None:
    if s is None or s == "":
        return None
    return float(s.replace(",", "."))


_PRE = {"atoll", "recif", "récif", "banc", "iles", "ile", "isla", "islas",
        "isle", "pulau", "kepulauan"}
_POST = {"atoll", "reef", "reefs", "bank", "banks", "shoal", "shoals",
         "islet", "islets", "island", "islands", "par", "atolls", "isle",
         "isles"}
_ARTICLES = {"de", "du", "des", "la", "le", "les", "d"}


def keyize(s: str) -> str:
    """Aggressive normalisation: deaccent, lowercase, drop common geo-suffixes
    in English / French / Indonesian to maximise cross-file matching."""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[\.,;:\-_'`\"()/&]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    parts = s.split()
    while parts and parts[0] in _PRE:
        parts.pop(0)
        while parts and parts[0] in _ARTICLES:
            parts.pop(0)
    while parts and parts[-1] in _POST:
        parts.pop()
    return " ".join(parts)


def load_latlon() -> dict[str, dict]:
    out: dict[str, dict] = {}
    with LATLON.open(encoding="utf-8") as f:
        rdr = csv.reader(f, delimiter="\t")
        next(rdr)
        for row in rdr:
            if not row:
                continue
            atoll, pays, lon, lat = row[0], row[1], row[2], row[3]
            out[atoll] = {
                "name": atoll,
                "pays_csv": pays,
                "lon": _f(lon),
                "lat": _f(lat),
                "key": keyize(atoll),
            }
    return out


def match_latlon(stat_name: str, latlon: dict[str, dict],
                 used: set[str]) -> dict | None:
    """Three-tier match: exact → keyized → fuzzy on unused entries."""
    # tier 1: exact
    if stat_name in latlon and stat_name not in used:
        return latlon[stat_name]
    # tier 2: keyized exact (1 candidate, not yet used)
    k = keyize(stat_name)
    candidates = [v for v in latlon.values()
                  if v["key"] == k and v["name"] not in used]
    if len(candidates) == 1:
        return candidates[0]
    # tier 3: fuzzy on unused keys
    unused = [v for v in latlon.values() if v["name"] not in used]
    keys = [v["key"] for v in unused]
    near = difflib.get_close_matches(k, keys, n=1, cutoff=0.6)
    if near:
        for v in unused:
            if v["key"] == near[0]:
                return v
    return None


def load_codes() -> dict[int, dict]:
    wb = openpyxl.load_workbook(CODES, data_only=True)
    ws = wb["Reefbase"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    out: dict[int, dict] = {}
    for r in rows[1:]:
        rec = dict(zip(header, r))
        code = rec.get("L5_CODE")
        if isinstance(code, int):
            out[code] = {
                "code": code,
                "l1": rec.get("L1_ATTRIB_N"),
                "l2": rec.get("L2_ATTRIB_N"),
                "l3": rec.get("L3_ATTRIB_N"),
                "l4": (rec.get("L4_ATTRIB_N") or "").strip(),
                "reef": bool(rec.get("REEF")),
                "depth": rec.get("DEPTH_ATTRIB"),
                "land": bool(rec.get("LAND")),
            }
    return out


def load_stats(latlon: dict[str, dict], codes: dict[int, dict]) -> list[dict]:
    wb = openpyxl.load_workbook(STATS, data_only=True)
    ws5 = wb["L5 km2"]
    ws4 = wb["L4 km2"]

    # L5 headers: [region, archipelago, atoll, code1, code2, ..., 'Total général']
    l5_hdr = [c.value for c in ws5[1]]
    l5_codes = [c for c in l5_hdr if isinstance(c, int)]
    total_idx = l5_hdr.index("Total général")

    # L4 headers: [region, archipelago, atoll, name1, name2, ...]
    l4_hdr = [c.value for c in ws4[1]]
    l4_classes = l4_hdr[3:]

    # Index L4 rows by atoll name
    l4_by_atoll: dict[str, dict] = {}
    for row in ws4.iter_rows(min_row=2, values_only=True):
        atoll = row[2]
        if atoll is None:
            continue
        scale = 1e-6 if atoll in L4_UNIT_FIX else 1.0
        breakdown = {}
        for cls, v in zip(l4_classes, row[3:]):
            if isinstance(v, (int, float)) and v:
                breakdown[cls] = round(v * scale, 4)
        l4_by_atoll[atoll] = breakdown

    out: list[dict] = []
    skipped_no_latlon: list[str] = []
    used: set[str] = set()
    for row in ws5.iter_rows(min_row=2, values_only=True):
        region, archipelago, atoll = row[0], row[1], row[2]
        if atoll is None:
            continue
        total = row[total_idx]
        l5_breakdown: dict[str, float] = {}
        for code, v in zip(l5_codes, row[3:total_idx]):
            if isinstance(v, (int, float)) and v:
                meta = codes.get(code, {})
                label = meta.get("l4") or f"code_{code}"
                # Several L5 codes share an L4 label; key by code to preserve detail
                l5_breakdown[str(code)] = {
                    "label": label,
                    "l3": meta.get("l3"),
                    "l2": meta.get("l2"),
                    "l1": meta.get("l1"),
                    "reef": meta.get("reef"),
                    "depth": meta.get("depth"),
                    "land": meta.get("land"),
                    "km2": round(v, 4),
                }
        ll = match_latlon(atoll, latlon, used) or {}
        if ll:
            used.add(ll["name"])
        else:
            skipped_no_latlon.append(atoll)
        rec = {
            "name": atoll,
            "region": region,
            "archipelago": archipelago,
            "pays_csv": ll.get("pays_csv"),
            "matched_latlon_name": ll.get("name"),
            "lat": ll.get("lat"),
            "lon": ll.get("lon"),
            "area_km2": round(total, 4) if isinstance(total, (int, float)) else None,
            "l5_classes_n": len(l5_breakdown),
            "l4": l4_by_atoll.get(atoll, {}),
            "l5": l5_breakdown,
        }
        out.append(rec)

    if skipped_no_latlon:
        print(f"[warn] {len(skipped_no_latlon)} atolls missing lat/lon: "
              f"{skipped_no_latlon[:5]}…")
    return out


def main() -> None:
    latlon = load_latlon()
    codes = load_codes()
    atolls = load_stats(latlon, codes)
    atolls.sort(key=lambda a: -(a["area_km2"] or 0))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT_PROCESSED.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "source": "Andréfouët S. & Paul M. 2023, MCRMP, DataSuds DOI 10.23708/OKTEFB",
        "license": "CC-BY-NC-SA 4.0",
        "n": len(atolls),
        "notes": [
            "L4 sheet values for 'Taka Garlarang' and 'Cato Reef' divided by 1e6 "
            "(stored in m² instead of km² in source).",
            "area_km2 is L5 'Total général'; l4{} is per named class; "
            "l5{} is per L5 code with codes lookup expansion.",
        ],
        "atolls": atolls,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")))
    OUT_PROCESSED.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
    print(f"[ok] wrote {len(atolls)} atolls → {OUT} "
          f"({OUT.stat().st_size/1024:.1f} KB)")
    print(f"[ok] pretty copy → {OUT_PROCESSED}")
    print(f"[stats] regions:",
          sorted({a['region'] for a in atolls if a['region']}))
    print(f"[stats] top 3 by area:",
          [(a['name'], a['area_km2']) for a in atolls[:3]])


if __name__ == "__main__":
    main()
